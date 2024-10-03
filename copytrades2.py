import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
import logging
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

wb = Workbook()
currency_style = NamedStyle(name="currency_style")

# Setup logging configuration
logging.basicConfig(
    filename='parsing_log.log',  # Log to file
    level=logging.DEBUG,  # Log level
    format='%(asctime)s - %(levelname)s - %(message)s'  # Log format
)

# Log to console as well
console = logging.StreamHandler()  # Create a console handler
console.setLevel(logging.INFO)  # Set logging level for console
console.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))  # Optional: set format for console logs

# Add the console handler to the root logger
logging.getLogger().addHandler(console)

# Initialize the workbook
wb = Workbook()

# Check if the style already exists in the workbook before adding
if "number_style_plus" not in wb.named_styles:
    number_style_p = NamedStyle(name="number_style_p", number_format='0.00000000')
    wb.add_named_style(number_style_p)

if "currency_style_red_black" not in wb.named_styles:
    currency_style_red_black = NamedStyle(name="currency_style_red_black", number_format='"$"#,##0.00_);[Red]"$"#,##0.00')
    wb.add_named_style(currency_style_red_black)

# Create named styles for formatting
currency_style = NamedStyle(name="currency_style", number_format='$#,##0.00')
number_style_p = NamedStyle(name="number_style_p", number_format='0.00000000')
currency_style_red_black = NamedStyle(name="currency_style_red_black", number_format='"$"#,##0.00_);[Red]"$"#,##0.00')


def extract_pdf_data_with_pdfplumber(pdf_path):
    logging.info(f"Extracting data from {pdf_path} using pdfplumber")
    trades = []
    columns = ["Symbol & Name", "Cusip", "Trade Date", "Settlement Date", "Account Type", "Buy/Sell", "Quantity", "Price", "Gross Amount", "Commission", "Fee/Tax", "Net Amount", "MKT", "Solicitation", "CAP"]

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages[2:], start=3):  # Start from the third page
                logging.info(f"Processing page {page_num}")
                tables = page.extract_tables()
                for table_num, table in enumerate(tables, start=1):
                    logging.info(f"Processing table {table_num} on page {page_num}")
                    if len(table[0]) == len(columns):
                        logging.info(f"Table {table_num} on page {page_num} matches expected column structure")
                        for row in table[1:]:  # Skip the header row
                            trade = {columns[i]: str(cell).strip() for i, cell in enumerate(row)}
                            trades.append(trade)
                            logging.info(f"Captured trade data: {trade}")
                    else:
                        logging.warning(f"Table {table_num} on page {page_num} does not match expected column structure")

    except Exception as e:
        logging.error(f"Error processing PDF {pdf_path}: {str(e)}")

    logging.info(f"Extracted {len(trades)} trades from {pdf_path}")
    return trades

def find_last_row(worksheet):

    last_row = worksheet.max_row
    while last_row > 0:
        if worksheet.cell(row=last_row, column=1).value is not None:  # Assuming data is in column 1
            break
        last_row -= 1
    return last_row

def append_trades_to_excel(trades, ws_trade_details, ws_trade_outcome):
    logging.info(f"Appending {len(trades)} trades to Excel sheets")
    
    existing_trades = set()
    last_row_trade_details = find_last_row(ws_trade_details)
    last_row_trade_outcome = find_last_row(ws_trade_outcome)

    for row in ws_trade_details.iter_rows(min_row=2, max_row=last_row_trade_details, values_only=True):
        if row[0]:  # Check if the row is not empty
            existing_trades.add((row[0], row[2], row[3], row[7]))  # Date, Symbol, Shares, Order Type

    new_trades_count = 0
    for trade in trades:
        trade_key = (
            trade['Trade Date'],
            trade['Symbol & Name'].split()[0],
            abs(float(trade['Quantity'].replace(',', ''))),
            'Buy' if trade['Buy/Sell'] == 'B' else 'Sell'
        )
        
        if trade_key not in existing_trades:
            # Mapping for Trade Entry Details
            trade_details_row = [
                trade['Trade Date'],
                '',  # Time (not available in your data)
                trade['Symbol & Name'].split()[0],  # Ticker Symbol
                abs(float(trade['Quantity'].replace(',', ''))),  # Shares (absolute value)
                abs(float(trade['Gross Amount'].replace(',', ''))),  # Position Size (absolute value)
                float(trade['Price'].replace(',', '')) if trade['Buy/Sell'] == 'B' else '',  # Entry Price
                float(trade['Price'].replace(',', '')) if trade['Buy/Sell'] == 'S' else '',  # Exit Price
                'Buy' if trade['Buy/Sell'] == 'B' else 'Sell',  # Order Type
                'Long'  # Assuming all trades are Long
            ]
            ws_trade_details.append(trade_details_row)
            logging.info(f"Appended new trade details: {trade_details_row}")
            
            next_row_trade_details = ws_trade_details.max_row + 1         


                     # Apply number style to the relevant cells
            for col in [4]:  
                ws_trade_details.cell(row=next_row_trade_details + new_trades_count, column=col).style = number_style_p

            
            # Apply currency style to the relevant cells
            for col in [5, 6, 7]:  # Assuming these columns need currency formatting
                ws_trade_details.cell(row=next_row_trade_details + new_trades_count, column=col).style = currency_style_red_black
                
                            # Apply currency style to the Trade Outcome cells
            for col in [2, 3, 4]:  # Columns B, C, D in Trade Outcome sheet
                ws_trade_outcome.cell(row=next_row_trade_outcome + new_trades_count, column=col).style = currency_style_red_black

            # Mapping for Trade Outcome
            outcome_row = [
                '',  # Empty cell for column A
                float(trade['Commission'].replace(',', '')),  # Commissions and Fees
                float(trade['Fee/Tax'].replace(',', '')),  # Tax
                float(trade['Net Amount'].replace(',', ''))  # Net
            ]
            ws_trade_outcome.append(outcome_row)
            logging.info(f"Appended new trade outcome: {outcome_row}")
            
            new_trades_count += 1
            existing_trades.add(trade_key)  # Add the new trade to existing trades
        else:
            logging.info(f"Skipped duplicate trade: {trade_key}")

    logging.info(f"Appended {new_trades_count} new trades to Excel sheets")

def get_or_create_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        logging.info(f"Creating new sheet: {sheet_name}")
        return wb.create_sheet(sheet_name)
    return wb[sheet_name]

def select_folder_or_file(prompt):
    root = tk.Tk()
    root.withdraw()
    
    if "folder" in prompt.lower():
        path = filedialog.askdirectory(title=prompt)
    else:
        path = filedialog.askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])
    
    return path

def main():
    print("Please select the folder containing the PDFs.")
    pdf_folder = select_folder_or_file("Select the folder containing the PDFs")
    
    print("Please select the Excel file.")
    excel_path = select_folder_or_file("Select the Excel file")

    logging.info(f"Starting to process PDF files in folder: {pdf_folder}")

    if not os.path.exists(pdf_folder):
        logging.error(f"The folder {pdf_folder} does not exist.")
        return
    if not os.path.exists(excel_path):
        logging.error(f"The file {excel_path} does not exist.")
        return

    try:
        wb = load_workbook(excel_path)
        logging.info(f"Successfully loaded Excel file: {excel_path}")
    except Exception as e:
        logging.error(f"Error opening Excel file {excel_path}: {str(e)}")
        return

    ws_trade_details = get_or_create_sheet(wb, 'Trade Entry Details')
    ws_trade_outcome = get_or_create_sheet(wb, 'Trade Outcome')

    pdf_files = [f for f in os.listdir(pdf_folder) if f.endswith('.pdf')]

    if not pdf_files:
        logging.warning(f"No PDF files found in the folder {pdf_folder}.")
    else:
        all_trades = []
        for pdf_file in pdf_files:
            pdf_path = os.path.join(pdf_folder, pdf_file)
            logging.info(f"Processing {pdf_file}...")
            trades = extract_pdf_data_with_pdfplumber(pdf_path)
            all_trades.extend(trades)

        if all_trades:
            append_trades_to_excel(all_trades, ws_trade_details, ws_trade_outcome)
        else:
            logging.warning("No trades found in any PDF.")

    wb.save(excel_path)
    logging.info(f"Workbook saved to {excel_path}")

    logging.info("Script execution completed.")

if __name__ == "__main__":
    main()
