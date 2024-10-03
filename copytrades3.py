import os
import pdfplumber
import logging
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import threading
import time
import csv
import sys
import shutil
from tkinter import ttk
import winsound
from tkinter import messagebox



# Determine if the app is running as a PyInstaller bundle
if getattr(sys, 'frozen', False):
    # If running as a PyInstaller bundle
    base_path = sys._MEIPASS
else:
    # If running in a normal Python environment
    base_path = os.path.dirname(__file__)

# Construct the full path to the bundled file
lic_path = os.path.join(base_path, 'lic')
readme_path = os.path.join(base_path, 'README.md')

def setup_logging():
    log_file = 'parsing_log.log'
    
    # Remove the old log file if it exists
    if os.path.exists(log_file):
        os.remove(log_file)
    
    logging.basicConfig(
        filename=log_file,
        filemode='w',  # 'w' mode overwrites the file
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    # Log to console as well
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logging.getLogger('').addHandler(console)

# Call this function at the start of your script
setup_logging()

def close_splash_screen(splash, root):
    splash.destroy()
    root.deiconify()

def show_splash_screen(root):
    splash = tk.Toplevel()
    splash.geometry("300x150")
    splash.title("Loading...")
    label = tk.Label(splash, text="App is loading...", font=("Helvetica", 16))
    label.pack(expand=True)
    root.withdraw()
    root.after(3000, close_splash_screen, splash, root)

def get_last_row(ws):
    max_row = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
    return max_row


def select_folder_or_file(prompt, select_file=False):
    root = tk.Tk()
    root.withdraw()

    if select_file:
        path = filedialog.askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])
    else:
        path = filedialog.askdirectory(title=prompt)

    return path

def extract_pdf_data_with_pdfplumber(pdf_path):
    logging.info(f"Extracting data from {pdf_path} using pdfplumber")
    trades = []
    columns = [
        "Symbol & Name", "Cusip", "Trade Date", "Settlement Date", "Account Type", "Buy/Sell",
        "Quantity", "Price", "Gross Amount", "Commission", "Fee/Tax", "Net Amount", "MKT", "Solicitation", "CAP"
    ]

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages[2:], start=3):
                logging.info(f"Processing page {page_num}")
                tables = page.extract_tables()
                for table_num, table in enumerate(tables, start=1):
                    logging.info(f"Processing table {table_num} on page {page_num}")
                    if len(table) > 0 and len(table[0]) == len(columns):
                        logging.info(f"Table {table_num} on page {page_num} matches expected column structure")
                        for row in table[1:]:
                            trade = {columns[i]: str(cell).strip() for i, cell in enumerate(row)}
                            trades.append(trade)
                            logging.info(f"Captured trade data: {trade}")
                    else:
                        logging.warning(f"Table {table_num} on page {page_num} does not match expected column structure")

    except Exception as e:
        logging.error(f"Error processing PDF {pdf_path}: {str(e)}")

    logging.info(f"Extracted {len(trades)} trades from {pdf_path}")
    return trades



def write_trades_to_excel(trades, excel_path):
    
    global all_trades  # Ensures that all_trades is shared
    all_trades = []  # Initialize the list at the start of processing
    
    logging.info(f"Writing {len(trades)} trades to Excel file: {excel_path}")

    # Load existing workbook
    workbook = load_workbook(excel_path)
    
    # Access 'Trade Entry Details' sheet
    if 'Trade Entry Details' in workbook.sheetnames:
        ws_trade_details = workbook['Trade Entry Details']
    else:
        ws_trade_details = workbook.create_sheet('Trade Entry Details')
        ws_trade_details.append(['Date', 'Time', 'Ticker Symbol', 'Shares', 'Position Size', 'Entry Price', 'Exit Price', 'Order Type', 'Long/Short', 'Trade ID'])

    # Access 'Trade Outcome' sheet
    if 'Trade Outcome' in workbook.sheetnames:
        ws_trade_outcome = workbook['Trade Outcome']
    else:
        ws_trade_outcome = workbook.create_sheet('Trade Outcome')
        ws_trade_outcome.append(['Profit/Loss', 'Commissions and Fees', 'Tax', 'Net', 'Trade ID'])

    # Define the formats directly
    number_format_8_dec = '0.00000000'
    currency_format_red_black = '"$"#,##0.00_);[Red]"$"#,##0.00'

    # Get existing trades
    existing_trades = set()
    for row in ws_trade_details.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2] and row[3] and row[9]:  # Date, Symbol, Shares, Trade ID
            existing_trades.add((row[0], row[2], row[3], row[9]))

    # Write trade details and outcomes, grouped by CUSIP (which is mapped to Trade ID)
    new_trades_count = 0
    for trade in trades:
        trade_date = trade['Trade Date']
        symbol = trade['Symbol & Name'].split()[0]
        shares = float(trade['Quantity'].replace(',', ''))
        cusip = trade['Cusip']
        
        # Check if trade already exists
        if (trade_date, symbol, shares, cusip) in existing_trades:
            logging.info(f"Skipping existing trade: Date={trade_date}, Symbol={symbol}, Shares={shares}, CUSIP={cusip}")
            continue

        def find_last_row(worksheet):
            for row in range(worksheet.max_row, 0, -1):
                if any(cell.value for cell in worksheet[row]):
                    return row
            return 0

        last_row_details = find_last_row(ws_trade_details)
        last_row_outcome = find_last_row(ws_trade_outcome)


        ws_trade_details.append([
            trade_date,
            '',  # Time (not available in your data)
            symbol,
            shares,
            float(trade['Gross Amount'].replace(',', '')),
            float(trade['Price'].replace(',', '')) if trade['Buy/Sell'] == 'B' else '',
            float(trade['Price'].replace(',', '')) if trade['Buy/Sell'] == 'S' else '',
            'Buy' if trade['Buy/Sell'] == 'B' else 'Sell',
            'Long',
            cusip
        ])
        
        # Apply number format to relevant columns in 'Trade Entry Details'
        ws_trade_details.cell(row=ws_trade_details.max_row, column=4).number_format = number_format_8_dec
        
        for col in [5, 6, 7]:
            last_row_details = ws_trade_details.max_row
            ws_trade_details.cell(row=last_row_details, column=col).number_format = currency_format_red_black



        ws_trade_outcome.append([
            '',  # Profit/Loss is not calculated here, can be calculated later if needed
            float(trade['Commission'].replace(',', '')),
            float(trade['Fee/Tax'].replace(',', '')),
            float(trade['Net Amount'].replace(',', '')),
            cusip
        ])
        
        # Apply red/black currency format to columns in 'Trade Outcome'
        for col in [2, 3, 4]:
            last_row_outcome = ws_trade_outcome.max_row
            ws_trade_outcome.cell(row=last_row_outcome, column=col).number_format = currency_format_red_black


        # Add to existing trades set
        existing_trades.add((trade_date, symbol, shares, cusip))

    # Save workbook
    workbook.save(excel_path)
    logging.info(f"Excel file saved: {excel_path}")
    logging.info(f"Added {new_trades_count} new trades")

def process_files_and_update_progress(pdf_folder, excel_path, progress_bar, stop_event, root):
    global all_trades  # Ensures that all_trades is shared
    all_trades = []  # Initialize the list at the start of processing
    
    """
    This function will process the files and update the progress bar after each file.
    """
    pdf_files = [f for f in os.listdir(pdf_folder) if f.endswith('.pdf')]

    total_files = len(pdf_files)
    if total_files == 0:
        logging.warning(f"No PDF files found in the folder {pdf_folder}.")
        root.event_generate('<<ProcessingComplete>>', when='tail')
        return

    all_trades = []

    for i, pdf_file in enumerate(pdf_files, start=1):
        if stop_event.is_set():  # Check if the process should be stopped
            break

        pdf_path = os.path.join(pdf_folder, pdf_file)
        logging.info(f"Processing {pdf_file}...")
        trades = extract_pdf_data_with_pdfplumber(pdf_path)
        all_trades.extend(trades)

        # Update the progress bar after processing each PDF
        progress_percentage = (i / total_files) * 100
        progress_bar['value'] = progress_percentage
        root.update_idletasks()  # Force update of the progress bar

    if all_trades and not stop_event.is_set():
        write_trades_to_excel(all_trades, excel_path)
    else:
        logging.warning("No trades found in any PDF or the process was stopped.")
    
    # Schedule the completion event for the main thread
    root.after(100, lambda: complete_processing(root))

def start_processing_task(pdf_folder, excel_path, progress_bar, stop_event, root):
    """
    Start the task to process PDFs and update the Excel sheet in a separate thread.
    This will keep the UI responsive while updating the progress bar.
    """
    def run_task():
        process_files_and_update_progress(pdf_folder, excel_path, progress_bar, stop_event, root)

    # Run the task in a separate thread to avoid blocking the main thread (UI)
    task_thread = threading.Thread(target=run_task)
    task_thread.start()


def complete_processing(root):
    try:
        winsound.PlaySound('C:\\Windows\\Media\\chimes.wav', winsound.SND_FILENAME)
    except RuntimeError as e:
        print(f"Error playing sound: {e}")
    finally:
        root.event_generate('<<ProcessingComplete>>', when='tail')



def main_app(pdf_folder, excel_path):
    """
    Main app to set up the UI and start the long-running task in a separate thread.
    """
    root = tk.Tk()
    root.title("Processing Trades")
    root.attributes('-topmost', True)  # This makes the window always on top
    root.after(100, lambda: root.attributes('-topmost', False))

    # Create a progress bar
    progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
    progress_bar.pack(pady=40)

    # Stop event to interrupt the task if needed
    stop_event = threading.Event()

    def on_processing_complete(event):
        root.quit()  # This will break the mainloop and allow the script to continue

    root.bind('<<ProcessingComplete>>', on_processing_complete)

    # Start processing task after the window is fully initialized and shown
    root.after(100, lambda: start_processing_task(pdf_folder, excel_path, progress_bar, stop_event, root))

    # Start the Tkinter event loop (this ensures the window is displayed)
    root.mainloop()

    # After mainloop exits, destroy the window
    root.destroy()

    logging.info("Processing completed.")

   
    
def main():
    global all_trades  # Declaring it as a global variable
    all_trades = []  # Initialize an empty list to store all trades
    
    print("Please select the folder containing the PDFs.")
    pdf_folder = select_folder_or_file("Select the folder containing the PDFs")

    print("Please select the existing Excel file to update.")
    excel_path = select_folder_or_file("Select the existing Excel file to update", select_file=True)

    if not excel_path:
        logging.error("No Excel file selected. Exiting.")
        return

    logging.info(f"Starting to process PDF files in folder: {pdf_folder}")

    if not os.path.exists(pdf_folder):
        logging.error(f"The folder {pdf_folder} does not exist.")
        return

    # Launch the Tkinter app with progress bar
    main_app(pdf_folder, excel_path)
    if all_trades and not stop_event.is_set():
        write_trades_to_excel(all_trades, excel_path)
    else:
        logging.warning("No trades found in any PDF or the process was stopped.")


if __name__ == "__main__":
    main()