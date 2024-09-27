import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook, Workbook
import logging
import sys
import tkinter as tk
from tkinter import filedialog

# Setup logging configuration
logging.basicConfig(
    filename='parsing_log.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Also log to console
console = logging.StreamHandler()
console.setLevel(logging.INFO)
logging.getLogger('').addHandler(console)

# Mapping for MKT (Market Abbreviation to Full Description)
MKT_MAPPING = {
    "NYSE": "New York Stock Exchange",
    "USE": "Other US Exchange",
    "OTH": "Other",
    "OP": "Options",
    "OTC": "Over-the-Counter/NASDAQ",
    "MUTF": "Mutual Funds",
    "FOREX": "Foreign Exchange",
    "BTCX": "Buy to Close Cancel",
    "STOX": "Sell to Open Cancel",
    # Add more mappings as necessary
}

# Function to extract data from PDF using pdfplumber
def extract_pdf_data_with_pdfplumber(pdf_path):
    logging.info(f"Extracting data from {pdf_path} using PDFplumber")
    trades = []
    trading_activity_started = False
    headers = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    logging.warning(f"No text extracted from page {page_num} of {pdf_path}")
                    continue

                lines = text.split('\n')
                for line in lines:
                    if "SECURITIES TRADING ACTIVITY" in line:
                        logging.info(f"Found 'SECURITIES TRADING ACTIVITY' section on page {page_num}")
                        trading_activity_started = True
                        continue

                    if not trading_activity_started:
                        continue

                    if "TRADING SUMMARY" in line:
                        trading_activity_started = False
                        break

                    if "Symbol & Name" in line:
                        headers = line.split()
                        logging.info(f"Found headers: {headers}")
                        continue

                    trade_data = line.split()
                    if len(trade_data) >= len(headers):
                        try:
                            trade = {}
                            for i, header in enumerate(headers):
                                if header in ["Quantity", "Price", "Gross Amount", "Commission", "Fee/Tax", "Net Amount"]:
                                    trade[header] = float(trade_data[i].replace(',', ''))
                                elif header == "MKT":
                                    trade[header] = MKT_MAPPING.get(trade_data[i], trade_data[i])  # Map MKT to full description
                                else:
                                    trade[header] = trade_data[i]
                            
                            # Handle multi-word company names
                            if len(trade_data) > len(headers):
                                trade["Symbol & Name"] = " ".join(trade_data[:len(trade_data) - len(headers) + 1])
                            
                            trades.append(trade)
                            logging.info(f"Parsed trade: {trade}")
                        except Exception as e:
                            logging.error(f"Error parsing trade data: {e}")
                            logging.debug(f"Problematic line: {line}")
        
        logging.info(f"Extracted {len(trades)} trades from {pdf_path}")
        return trades

    except Exception as e:
        logging.error(f"Error processing PDF {pdf_path}: {e}")
        return []

# Function to get or create a sheet in the workbook
def get_or_create_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        logging.info(f"Creating new sheet: {sheet_name}")
        return wb.create_sheet(sheet_name)
    return wb[sheet_name]

# Function to append trades to the Excel sheet
def append_trades_to_excel(trades, ws_trade_details, ws_trade_outcome):
    logging.info(f"Appending {len(trades)} trades to Excel sheets")

    for trade in trades:
        row = [
            trade.get('Trade Date', ''),
            '',  # Time (not available)
            trade.get('Symbol & Name', '').split()[0],  # Only the symbol
            trade.get('Quantity', ''),
            trade.get('Price', ''),
            '',  # Exit Price (not available)
            trade.get('Buy/Sell', ''),
            'Long',  # Assume all trades are Long
            trade.get('MKT', '')  # Add MKT (market) column
        ]
        ws_trade_details.append(row)
        logging.info(f"Appended trade details: {row}")

        # Append trade outcome (Net Amount and Commission)
        outcome_row = [
            trade.get('Net Amount', 0),
            trade.get('Commission', 0)  # Use actual commission if present, otherwise default to 0
        ]
        ws_trade_outcome.append(outcome_row)
        logging.info(f"Appended trade outcome: {outcome_row}")

# Function to select a folder or file using a dialog window
def select_folder_or_file(prompt):
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    if "folder" in prompt.lower():
        path = filedialog.askdirectory(title=prompt)
    else:
        path = filedialog.askopenfilename(title=prompt, filetypes=[("Excel files", "*.xlsx")])
    
    return path

# Main function to process PDF files and update the Excel workbook
def main():
    print("Please select the folder containing the PDFs.")
    pdf_folder = select_folder_or_file("Select the folder containing the PDFs")
    
    print("Please select the Excel file.")
    excel_path = select_folder_or_file("Select the Excel file")

    logging.info(f"Starting to process PDF files in folder: {pdf_folder}")

    if not os.path.exists(pdf_folder):
        logging.error(f"The folder {pdf_folder} does not exist.")
        return
    if not os.path.exists(excel_path):
        logging.error(f"The file {excel_path} does not exist.")
        return

    try:
        wb = load_workbook(excel_path)
        logging.info(f"Successfully loaded Excel file: {excel_path}")
    except Exception as e:
        logging.error(f"Error opening Excel file {excel_path}: {e}")
        return

    ws_trade_details = get_or_create_sheet(wb, 'Trade Entry Details')
    ws_trade_outcome = get_or_create_sheet(wb, 'Trade Outcome')

    pdf_files = [f for f in os.listdir(pdf_folder) if f.endswith('.pdf')]

    if not pdf_files:
        logging.warning(f"No PDF files found in the folder {pdf_folder}.")
    else:
        all_trades = []
        for pdf_file in pdf_files:
            pdf_path = os.path.join(pdf_folder, pdf_file)
            logging.info(f"Processing {pdf_file}...")
            trades = extract_pdf_data_with_pdfplumber(pdf_path)
            all_trades.extend(trades)

        if all_trades:
            append_trades_to_excel(all_trades, ws_trade_details, ws_trade_outcome)
        else:
            logging.warning("No trades found in any PDF.")

    wb.save(excel_path)
    logging.info(f"Workbook saved to {excel_path}")

    logging.info("Script execution completed.")

if __name__ == "__main__":
    main()
